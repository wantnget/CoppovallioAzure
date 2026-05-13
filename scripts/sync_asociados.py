import os
import sys
import json
import math
import datetime
import unicodedata
import urllib.request
import urllib.error
import openpyxl

EXCEL_PATH   = "Simulador_Analisis_Creditos_V3 2025.xlsx"
SHEET_NAME   = "Base_Asociados"
HEADER_ROW   = 5
DATA_START   = 6

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
TABLE_NAME   = "datos_asociado"
BATCH_SIZE   = 200

COLUMN_MAP = {
    "Cedula":                           "cedula",
    "apellido":                         "primer_apellido",
    "nombre":                           "nombre",
    "ciudad":                           "ciudad",
    "estado civil":                     "estado_civil",
    "Salario":                          "salario",
    "aportes":                          "aportes",
    "deuda Coopvalili":                 "deuda_coopvalili",
    "edad":                             "edad",
    "Tipo Vivienda":                    "tipo_vivienda",
    "Fecha Ing Coopvalili":             "fecha_ingreso",
    "Fecha Ing FVL":                    "fecha_ingreso_empresa",
    "personas a cargo":                 "personas_cargo",
    "Empresa":                          "cliente_empresa",
    "Nivel":                            "nivel",
    "Cuota Disponible FVL":             "cuota_disponible",
    "ASOCIADO":                         "nombre_asociado",
    "Antigüedad en la Cooperativa":     "antiguedad_coop",
    "Antigüedad Laboral":               "antiguedad_laboral",
    "Usuario De Credito en Coopvalili": "usuario_credito",
    "ESTADO_CIVIL":                     "estado_civil_norm",
}

def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )

_COLUMN_MAP_NORM = {strip_accents(k): v for k, v in COLUMN_MAP.items()}

def clean_value(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("#N/A", "N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!"):
            return None
        return s if s != "" else None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, datetime.time):
        return v.isoformat()
    return v

def supabase_request(url: str, method: str = "GET", body: bytes = None) -> any:
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
    }
    req = urllib.request.Request(url, data=body, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"HTTP {e.code}: {e.read().decode()}") from e

def fetch_existing_cedulas() -> set:
    cedulas = set()
    page    = 0
    limit   = 1000
    while True:
        url  = f"{SUPABASE_URL}/rest/v1/{TABLE_NAME}?select=cedula&limit={limit}&offset={page * limit}"
        rows = supabase_request(url)
        if not rows:
            break
        for r in rows:
            cedulas.add(r["cedula"])
        if len(rows) < limit:
            break
        page += 1
    return cedulas

def supabase_insert(rows: list[dict]):
    url     = f"{SUPABASE_URL}/rest/v1/{TABLE_NAME}"
    payload = json.dumps(rows).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        method="POST",
        headers={
            "Content-Type":  "application/json",
            "apikey":        SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Prefer":        "return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            if resp.getcode() not in (200, 201):
                raise RuntimeError(f"HTTP {resp.getcode()}: {resp.read().decode()}")
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"HTTP {e.code}: {e.read().decode()}") from e

def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: Faltan SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY", file=sys.stderr)
        sys.exit(1)

    # ── 1. Leer Excel ────────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Hoja '{SHEET_NAME}' no encontrada", file=sys.stderr)
        sys.exit(1)

    ws          = wb[SHEET_NAME]
    raw_headers = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]

    columns = []
    for idx, h in enumerate(raw_headers):
        if h is None:
            continue
        key_norm = strip_accents(str(h).strip())
        if key_norm in _COLUMN_MAP_NORM:
            columns.append((idx, _COLUMN_MAP_NORM[key_norm]))

    print(f"Columnas mapeadas: {len(columns)}")

    # Leer filas y deduplicar por cédula (última ocurrencia gana)
    seen    = {}
    skipped = 0

    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        row_dict = {field: clean_value(row[idx] if idx < len(row) else None) for idx, field in columns}
        cedula   = row_dict.get("cedula")
        if cedula is None:
            skipped += 1
            continue
        try:
            float(str(cedula))
        except (ValueError, TypeError):
            skipped += 1
            continue
        row_dict["cedula"] = str(cedula).split(".")[0]
        seen[row_dict["cedula"]] = row_dict

    wb.close()
    all_records = list(seen.values())
    print(f"Registros unicos en Excel: {len(all_records)} | Saltadas: {skipped}")

    # ── 2. Consultar cédulas existentes en Supabase ──────────────────────────────
    print("Consultando cedulas existentes en Supabase...")
    existing = fetch_existing_cedulas()
    print(f"Cedulas ya en Supabase: {len(existing)}")

    # ── 3. Filtrar solo las nuevas ───────────────────────────────────────────────
    new_records = [r for r in all_records if r["cedula"] not in existing]
    print(f"Cedulas nuevas a insertar: {len(new_records)}")

    if not new_records:
        print("Sin cedulas nuevas. Nada que insertar.")
        sys.exit(0)

    # ── 4. Insertar en batches ───────────────────────────────────────────────────
    total_batches = math.ceil(len(new_records) / BATCH_SIZE)
    print(f"Insertando en {total_batches} batch(es)...")

    for i in range(0, len(new_records), BATCH_SIZE):
        batch     = new_records[i : i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Batch {batch_num}/{total_batches} ({len(batch)} registros)...", end=" ")
        supabase_insert(batch)
        print("OK")

    print(f"Completado: {len(new_records)} nuevos registros insertados en '{TABLE_NAME}'")

if __name__ == "__main__":
    main()
